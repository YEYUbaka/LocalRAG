import hashlib
import logging
from pathlib import Path

from langchain_community.document_loaders import (
    PyPDFLoader, Docx2txtLoader, TextLoader,
    UnstructuredPowerPointLoader,
    BSHTMLLoader, CSVLoader,
)
from langchain_core.documents import Document as LangChainDocument
from langchain_text_splitters import RecursiveCharacterTextSplitter
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.config import settings
from app.models import Document
from app.core.vectorstore import add_documents, delete_by_document_id
from app.domain.tenant import TenantScope

logger = logging.getLogger(__name__)

CHUNKER_VERSION = "1"


class OpenpyxlExcelLoader:
    """Load native XLSX tables with the project's existing openpyxl dependency."""

    def __init__(self, file_path: str):
        self.file_path = file_path

    def load(self) -> list[LangChainDocument]:
        workbook = load_workbook(self.file_path, read_only=True, data_only=True)
        try:
            documents = []
            for worksheet in workbook.worksheets:
                rows = []
                for values in worksheet.iter_rows(values_only=True):
                    if not any(value is not None for value in values):
                        continue
                    rows.append("\t".join("" if value is None else str(value) for value in values))
                if rows:
                    documents.append(LangChainDocument(
                        page_content="\n".join(rows),
                        metadata={"source": self.file_path, "sheet": worksheet.title},
                    ))
            return documents
        finally:
            workbook.close()


LOADER_MAP = {
    ".pdf": PyPDFLoader,
    ".docx": Docx2txtLoader,
    ".doc": Docx2txtLoader,
    ".md": TextLoader,
    ".txt": TextLoader,
    ".xlsx": OpenpyxlExcelLoader,
    ".pptx": UnstructuredPowerPointLoader,
    ".html": BSHTMLLoader,
    ".htm": BSHTMLLoader,
    ".csv": CSVLoader,
}


def compute_md5(file_path: Path) -> str:
    h = hashlib.md5()
    with open(file_path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def parse_document(file_path: Path) -> list:
    suffix = file_path.suffix.lower()
    loader_cls = LOADER_MAP.get(suffix)
    if not loader_cls:
        raise ValueError(f"不支持的文件格式: {suffix}")

    kwargs = {}
    if suffix in (".md", ".txt", ".csv"):
        kwargs["encoding"] = "utf-8"

    loader = loader_cls(str(file_path), **kwargs)
    return loader.load()


def split_documents(docs: list, filename: str) -> tuple[list[str], list[dict]]:
    splitter = RecursiveCharacterTextSplitter(
        chunk_size=settings.chunk_size,
        chunk_overlap=settings.chunk_overlap,
    )
    chunks = splitter.split_documents(docs)
    texts = [chunk.page_content for chunk in chunks]
    metadatas = []
    for chunk in chunks:
        meta = {**chunk.metadata, "filename": filename}
        metadatas.append(meta)
    return texts, metadatas


def build_stable_chunk_metadata(
    document_key: str,
    document_version: int,
    texts: list[str],
    metadatas: list[dict],
    chunker_version: str = CHUNKER_VERSION,
) -> list[dict]:
    """Attach deterministic cross-index identity to chunks in reading order."""
    enriched = []
    for ordinal, (text, metadata) in enumerate(zip(texts, metadatas, strict=True)):
        chunk_id = (
            f"{document_key}-v{document_version}-c{chunker_version}-{ordinal:06d}"
        )
        enriched.append({
            **metadata,
            "chunk_id": chunk_id,
            "document_key": document_key,
            "document_version": document_version,
            "chunker_version": chunker_version,
            "content_hash": hashlib.sha256(text.encode("utf-8")).hexdigest(),
        })
    return enriched


def _stable_identity(doc: Document, content: str | None = None) -> tuple[str, int, str]:
    document_key = doc.document_key
    if not document_key:
        document_key = doc.md5_hash or hashlib.sha256((content or "").encode("utf-8")).hexdigest()
        doc.document_key = document_key
    return (
        document_key,
        doc.document_version or 1,
        doc.chunker_version or CHUNKER_VERSION,
    )


def compute_page_breaks(raw_docs: list) -> list[int] | None:
    """Calculate character offsets where each page starts (PDF only)."""
    if not raw_docs or "page" not in raw_docs[0].metadata:
        return None

    breaks = []
    current_page = None
    offset = 0
    for doc in raw_docs:
        page = doc.metadata.get("page")
        if page != current_page:
            breaks.append(offset)
            current_page = page
        offset += len(doc.page_content)
    return breaks


def process_document(doc_id: int, db_session_factory) -> None:
    """BackgroundTasks 回调：解析文档并入库"""
    db: Session = db_session_factory()
    try:
        doc = db.query(Document).filter(Document.id == doc_id).first()
        if not doc:
            return

        doc.status = "processing"
        db.commit()

        file_path = Path(doc.file_path)
        raw_docs = parse_document(file_path)

        # Save parsed content and page breaks
        doc.parsed_content = "\n\n".join(d.page_content for d in raw_docs)
        doc.page_breaks = compute_page_breaks(raw_docs)

        texts, metadatas = split_documents(raw_docs, doc.filename)
        document_key, document_version, chunker_version = _stable_identity(doc, doc.parsed_content)
        metadatas = build_stable_chunk_metadata(
            document_key, document_version, texts, metadatas, chunker_version,
        )
        doc.chunk_count = len(texts)

        scope = TenantScope(user_id=doc.user_id or 1, kb_id=doc.kb_id or 1)
        add_documents(scope, doc_id, texts, metadatas)

        # Sync BM25 index
        try:
            from app.core.bm25_search import add_document_chunks
            add_document_chunks(scope, doc_id, texts, metadatas=metadatas)
        except Exception as e:
            logger.warning(f"BM25 sync failed for doc {doc_id}: {e}")

        doc.status = "completed"
        db.commit()
    except Exception as e:
        doc = db.query(Document).filter(Document.id == doc_id).first()
        if doc:
            doc.status = "failed"
            doc.error_message = str(e)
            db.commit()
    finally:
        db.close()


async def process_url_import(doc_id: int, url: str):
    """处理单个 URL 的导入。"""
    from app.core.web_fetcher import fetch_single_url
    from app.core.bm25_search import add_document_chunks
    from app.main import SessionLocal

    db = SessionLocal()
    try:
        doc = db.query(Document).filter(Document.id == doc_id).first()
        if not doc:
            return
        doc.status = "processing"
        db.commit()

        lc_doc = await fetch_single_url(url)
        texts, metadatas = split_documents([lc_doc], doc.filename)
        document_key, document_version, chunker_version = _stable_identity(doc, lc_doc.page_content)
        metadatas = build_stable_chunk_metadata(
            document_key, document_version, texts, metadatas, chunker_version,
        )
        scope = TenantScope(user_id=doc.user_id or 1, kb_id=doc.kb_id or 1)
        add_documents(scope, doc_id, texts, metadatas)
        add_document_chunks(scope, doc_id, texts, metadatas)

        doc.parsed_content = lc_doc.page_content
        doc.chunk_count = len(texts)
        doc.status = "completed"
        doc.file_size = len(lc_doc.page_content.encode("utf-8"))
        db.commit()
    except Exception as e:
        doc = db.query(Document).filter(Document.id == doc_id).first()
        if doc:
            doc.status = "failed"
            doc.error_message = str(e)[:500]
            db.commit()
    finally:
        db.close()


async def process_crawl_import(doc_id: int, start_url: str, max_pages: int, max_depth: int):
    """处理整站爬取导入。"""
    from app.core.web_fetcher import crawl_site
    from app.core.bm25_search import add_document_chunks
    from app.main import SessionLocal

    db = SessionLocal()
    try:
        doc = db.query(Document).filter(Document.id == doc_id).first()
        if not doc:
            return
        doc.status = "processing"
        db.commit()

        lc_docs = await crawl_site(start_url, max_pages=max_pages, max_depth=max_depth)
        if not lc_docs:
            doc.status = "failed"
            doc.error_message = "未抓取到任何内容"
            db.commit()
            return

        all_text = "\n\n---\n\n".join(d.page_content for d in lc_docs)
        texts, metadatas = split_documents(lc_docs, doc.filename)
        document_key, document_version, chunker_version = _stable_identity(doc, all_text)
        metadatas = build_stable_chunk_metadata(
            document_key, document_version, texts, metadatas, chunker_version,
        )
        scope = TenantScope(user_id=doc.user_id or 1, kb_id=doc.kb_id or 1)
        add_documents(scope, doc_id, texts, metadatas)
        add_document_chunks(scope, doc_id, texts, metadatas)

        doc.parsed_content = all_text
        doc.chunk_count = len(texts)
        doc.status = "completed"
        doc.file_size = len(all_text.encode("utf-8"))
        db.commit()
    except Exception as e:
        doc = db.query(Document).filter(Document.id == doc_id).first()
        if doc:
            doc.status = "failed"
            doc.error_message = str(e)[:500]
            db.commit()
    finally:
        db.close()


def delete_document(doc_id: int, db: Session) -> None:
    doc = db.query(Document).filter(Document.id == doc_id).first()
    if not doc:
        raise ValueError(f"文档 {doc_id} 不存在")

    scope = TenantScope(user_id=doc.user_id or 1, kb_id=doc.kb_id or 1)
    delete_by_document_id(scope, doc_id)

    # Remove from BM25 index
    try:
        from app.core.bm25_search import remove_document
        remove_document(doc_id)
    except Exception:
        pass

    file_path = Path(doc.file_path)
    if file_path.exists():
        file_path.unlink()

    db.delete(doc)
    db.commit()
